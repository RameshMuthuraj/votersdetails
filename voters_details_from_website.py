import openpyxl

from selenium import webdriver

from selenium.webdriver.common.keys import Keys

from time import sleep 


browser = webdriver.Firefox()

#browser.get("http://domain.com/form.html")

sleep(10)
#making sure everything is loaded. giving some time for the webpage to load.



wb = openpyxl.load_workbook('/home/bhishan/Downloads/bhishan_Ids_List.xlsx', read_only=False)

#wb.get_sheet_names()

sheet = wb.worksheets[0]

#req_id = sheet.cell(row=1, column=2).value
#this is how we get the value in a cell in excel file



# when you need to automate, this helps. Basically just iterating over all the rows in the excel sheet

highest_row = sheet.get_highest_row()
#highest_row = sheet.max_row()

for i in range(2, highest_row):

    browser.get("http://104.211.231.134/Search/searchbyname.aspx")
    #get the webpage where you need to submit information. 

    req_id = sheet.cell(row = i, column = 1).value
    # i assume the id's are in the first column else replace with corresponding column name. Also rewrite the loop if it starts from row other than 1.

    input_for_id = browser.find_element_by_id('txt_Epicno')
    #inspect element to see the id associated with input field. Additionally if id is not available various methods such as find_element_by_class_name , find_element_by_xpath, etc methods are available. Google for em. 
    
    input_for_id.send_keys(req_id) # passing the id to the input field

    input_for_id.send_keys(Keys.RETURN) # submit the form

    sleep(7) # increase or decrease acc to trust to your internet speed. Always safe to keep the value more. 
    id_card_no = browser.find_element_by_id('lbepicen').text
    print id_card_no
    sheet.cell(row = i, column = 2)
    ac_no_name = browser.find_element_by_id('lbacnameen').text
    sheet.cell(row = i, column = 3).value = ac_no_name
    elector_name = browser.find_element_by_id('lbnameen').text
    sheet.cell(row = i, column = 4).value = elector_name
    relation_name = browser.find_element_by_id('lbrlnen').text
    sheet.cell(row = i, column = 5).value = relation_name
    gender = browser.find_element_by_id('lbsexen').text
    sheet.cell(row = i, column = 6).value = gender
    address = browser.find_element_by_id('lbaden').text
    sheet.cell(row = i, column = 7).value = address
    polling_station = browser.find_element_by_id('lbpartnameen').text
    sheet.cell(row = i, column = 8).value = polling_station
    part_no = browser.find_element_by_id('lbparten').text
    sheet.cell(row = i, column = 9).value = part_no
    serial_no = browser.find_element_by_id('lbserialen').text
    sheet.cell(row = i, column = 10).value = serial_no

    #info_from_website = browser.find_element_by_id("idnameofinfo").text


    
    # now once you submit the req_id to the website and receive the infomations from there, write to the right of the current cell.
    #sheet.cell(row = i, column = 2).value = info_from_website

    wb.save("/home/bhishan/Downloads/bhishan_Ids_List.xlsx")
#wb.save('/home/bhishan/Downloads/bhishan_Ids_List.xlsx')

